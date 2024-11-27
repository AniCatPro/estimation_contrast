import openpyxl
from openpyxl.styles import PatternFill
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox


def apply_style_to_row(ws, row_idx, style):
    """Применяет стиль ко всем ячейкам строки по индексу."""
    for cell in ws[row_idx]:
        if cell.value is not None:  # Проверяем, чтобы не было None
            cell.fill = style


def compare_header_and_footer(ws1, ws2):
    """Сравнивает шапку и итоговые строки (с начала до первого раздела и после разделов)."""
    header_footer_differences = []
    # Шапка - до первого раздела
    for row_idx, row1 in enumerate(ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column), 1):
        row2 = ws2[row_idx]
        if row1[0].value and isinstance(row1[0].value, str) and "Раздел" in row1[0].value:
            break  # Конец шапки, начинаются разделы
        for cell1, cell2 in zip(row1, row2):
            if cell1.value != cell2.value:
                header_footer_differences.append((row_idx, cell1.column, cell1.value, cell2.value))

    # Итоги по смете - от "Итоги по смете:" до конца
    for row_idx in range(ws1.max_row, 0, -1):
        row1 = ws1[row_idx]
        row2 = ws2[row_idx]
        if row1[0].value and isinstance(row1[0].value, str) and "Итоги по смете" in row1[0].value:
            break  # Начинаются итоги, конец сметы
        for cell1, cell2 in zip(row1, row2):
            if cell1.value != cell2.value:
                header_footer_differences.append((row_idx, cell1.column, cell1.value, cell2.value))

    return header_footer_differences


def compare_sections_and_works(ws1, ws2):
    """Сравнивает разделы и работы по принципу: добавлено, удалено или изменено."""
    section_differences = []
    current_section1 = None
    current_section2 = None

    # Стиль для подсветки изменений
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_idx, row1 in enumerate(ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column), 1):
        row2 = ws2[row_idx]
        # Проверка на строку с разделом (например, "Раздел 1. Срубка оголовков свай")
        if row1[0].value and isinstance(row1[0].value, str) and "Раздел" in row1[0].value:
            current_section1 = row1[0].value
            current_section2 = row2[0].value
        elif row1[0].value and isinstance(row1[0].value, str) and row1[0].value.isdigit():  # Это работа
            # Сравниваем только работы, если номер в колонке A
            work_data1 = [cell.value if cell.value is not None else "" for cell in row1[1:15]]  # Рабочие ячейки B:O
            work_data2 = [cell.value if cell.value is not None else "" for cell in row2[1:15]]  # Рабочие ячейки B:O
            if work_data1 != work_data2:
                section_differences.append((row_idx, work_data1, work_data2))
                apply_style_to_row(ws2, row_idx, yellow_fill)
    return section_differences


def compare_excel_files(file1, file2, save_path):
    # Открываем оба файла
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Копируем второй файл для внесения изменений
    wb2_copy = openpyxl.load_workbook(file2)

    # Стиль для подсветки изменений
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Список листов в обоих файлах
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames

    # Для каждого листа
    for i in range(min(len(sheets1), len(sheets2))):
        ws1 = wb1[sheets1[i]]
        ws2 = wb2_copy[sheets2[i]]

        # Сравниваем шапку и конец сметы
        header_footer_changes = compare_header_and_footer(ws1, ws2)

        # Подсвечиваем изменения в шапке и итогах
        for change in header_footer_changes:
            row_idx, col_idx, val1, val2 = change
            if val1 != val2:
                ws2.cell(row=row_idx, column=col_idx).fill = yellow_fill

        # Сравниваем разделы и работы
        section_changes = compare_sections_and_works(ws1, ws2)

        # Подсвечиваем изменения в работах
        for change in section_changes:
            row_idx, work1, work2 = change
            if work1 != work2:
                ws2.cell(row=row_idx, column=1).fill = yellow_fill  # Подсвечиваем всю строку работы

    # Формируем имя нового файла
    base_name = os.path.basename(file2)
    new_file_name = f"СРВ_{base_name}"
    new_file_path = os.path.join(save_path, new_file_name)

    # Сохраняем новый файл с подсветкой изменений
    wb2_copy.save(new_file_path)
    return new_file_path


class ExcelCompareApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Сравнение Excel файлов")

        # Кнопка для выбора первого файла
        self.file1_button = tk.Button(root, text="Выбрать старый файл", command=self.select_file1)
        self.file1_button.pack(pady=10)

        # Кнопка для выбора второго файла
        self.file2_button = tk.Button(root, text="Выбрать новый файл", command=self.select_file2)
        self.file2_button.pack(pady=10)

        # Кнопка для выбора папки сохранения
        self.save_button = tk.Button(root, text="Выбрать папку для сохранения", command=self.select_save_path)
        self.save_button.pack(pady=10)

        # Кнопка для сравнения файлов
        self.compare_button = tk.Button(root, text="Сравнить файлы", command=self.compare_files)
        self.compare_button.pack(pady=20)

        # Метки для отображения выбранных файлов и папки
        self.file1_label = tk.Label(root, text="Старый файл: не выбран")
        self.file1_label.pack(pady=5)

        self.file2_label = tk.Label(root, text="Новый файл: не выбран")
        self.file2_label.pack(pady=5)

        self.save_label = tk.Label(root, text="Папка для сохранения: не выбрана")
        self.save_label.pack(pady=5)

        # Переменные для хранения путей файлов и папки
        self.file1_path = ""
        self.file2_path = ""
        self.save_path = ""

    def select_file1(self):
        self.file1_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.file1_label.config(text=f"Старый файл: {self.file1_path if self.file1_path else 'не выбран'}")

    def select_file2(self):
        self.file2_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.file2_label.config(text=f"Новый файл: {self.file2_path if self.file2_path else 'не выбран'}")

    def select_save_path(self):
        self.save_path = filedialog.askdirectory()
        self.save_label.config(text=f"Папка для сохранения: {self.save_path if self.save_path else 'не выбрана'}")

    def compare_files(self):
        if not all([self.file1_path, self.file2_path, self.save_path]):
            messagebox.showerror("Ошибка", "Пожалуйста, выберите все файлы и папку для сохранения!")
            return

        try:
            new_file = compare_excel_files(self.file1_path, self.file2_path, self.save_path)
            messagebox.showinfo("Готово", f"Сравнение завершено! Новый файл сохранен по пути: {new_file}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при сравнении: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelCompareApp(root)
    root.mainloop()
